"""Build a single, ready-to-complete formal IAA XLSX for 12 papers."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import shutil
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_HUMAN_PATH = (
    "outputs/validation_sampling/validation_sampling_v0_2_seed20260717_main_n48.csv"
)
DEFAULT_IAA_PATH = (
    "outputs/validation_sampling/"
    "validation_sampling_v0_2_seed20260717_formal_iaa_human_n12.csv"
)
DEFAULT_LLM_PATH = (
    "outputs/review_generation/analysis/"
    "validation_sampling_v0_2_seed20260717_main_n48_"
    "zero_shot_review_generation_v0_6_pdf_n48_requests_results_"
    "gpt-5.4_temp0.0_reviews.csv"
)
DEFAULT_CHECKLIST_PATH = "artifacts/checklist.csv"
DEFAULT_GUIDELINE_PATH = "artifacts/annotation_guideline.md"
DEFAULT_OUTPUT_DIR = "outputs/review_generation/formal_iaa"
DEFAULT_WORKBOOK_NAME = "formal_iaa_annotation_master_v0_1.xlsx"
DEFAULT_ANNOTATOR_GUIDELINE_NAME = "formal_iaa_guideline_v0_1.md"

ANSWER_OPTIONS = ("yes", "partial", "no", "not_applicable")
TABLE_HEADER_ROW = 24
FIRST_QUESTION_ROW = 25
LAST_QUESTION_ROW = 49

COLOURS = {
    "primary": "244A57",
    "primary_light": "DDE9EC",
    "human": "DCEAF7",
    "llm": "FCE8D5",
    "input": "FFF4CC",
    "header": "E7EBEF",
    "stripe": "F7F8FA",
    "text": "243038",
    "muted": "59656D",
    "border": "C8D0D6",
    "yes": "D9EAD3",
    "partial": "FFF2CC",
    "no": "F4CCCC",
    "na": "E7E6E6",
}

SECTION_HEADING_MAP = {
    "paper_summary": "Summary",
    "summary_of_strengths": "Strengths",
    "summary_of_weaknesses": "Weaknesses",
    "comments,_suggestions_and_typos": "Comments and suggestions",
    "comments_suggestions_and_typos": "Comments and suggestions",
    "comments_suggestions_and_questions": "Comments and suggestions",
    "ethical_concerns": "Ethical concerns",
}


def resolve_path(path_value: str | Path) -> Path:
    path = Path(path_value)
    return path if path.is_absolute() else PROJECT_ROOT / path


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def as_bool(value: str) -> bool:
    return value.strip().lower() in {"true", "1", "yes"}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as f:
        for block in iter(lambda: f.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def normalise_section_headings(review_text: str) -> str:
    """Normalize display headings and add an empty ethics section.

    Leave the review body unchanged.
    """

    for original, display in SECTION_HEADING_MAP.items():
        review_text = review_text.replace(f"[{original}]", f"[{display}]")
    if "[Ethical concerns]" not in review_text:
        review_text = review_text.rstrip() + "\n\n[Ethical concerns]\n"
    return review_text


def compact_markdown_text(text: str) -> str:
    return " ".join(line.strip() for line in text.strip().splitlines() if line.strip())


def parse_item_guideline(guideline_text: str, question_id: str) -> dict[str, str]:
    section_match = re.search(
        rf"^### {re.escape(question_id)}\s*$\n(.*?)(?=^### |\Z)",
        guideline_text,
        flags=re.MULTILINE | re.DOTALL,
    )
    if not section_match:
        raise ValueError(f"Guideline is missing item section: {question_id}")

    section = section_match.group(1)
    labels = [
        "Question",
        "Context required",
        "Definition",
        "Yes",
        "Partial",
        "No",
        "Not applicable",
        "Boundary note",
    ]
    parsed: dict[str, str] = {}
    for label in labels:
        value_match = re.search(
            rf"\*\*{re.escape(label)}:\*\*\s*(.*?)(?=\n\n\*\*|\Z)",
            section,
            flags=re.DOTALL,
        )
        parsed[label] = compact_markdown_text(value_match.group(1)) if value_match else ""
    return parsed


def parse_pairwise_boundaries(guideline_text: str) -> list[list[str]]:
    section_match = re.search(
        r"^## Pairwise Boundary Rules\s*$\n(.*?)(?=^## Checklist Items)",
        guideline_text,
        flags=re.MULTILINE | re.DOTALL,
    )
    if not section_match:
        return []

    rows: list[list[str]] = []
    for line in section_match.group(1).splitlines():
        if not line.startswith("|") or "---" in line or "Main distinction" in line:
            continue
        cells = [cell.strip().replace("`", "") for cell in line.strip("|").split("|")]
        if len(cells) == 3:
            rows.append(cells)
    return rows


def build_paper_records(
    human_rows: list[dict[str, str]],
    iaa_rows: list[dict[str, str]],
    llm_rows: list[dict[str, str]],
    paper_dir: Path,
) -> list[dict[str, Any]]:
    human_by_pair = {row["pair_id"]: row for row in human_rows}
    llm_by_pair = {row["pair_id"]: row for row in llm_rows}
    iaa_pair_ids = {row["pair_id"] for row in iaa_rows}
    if len(iaa_pair_ids) != 12 or not iaa_pair_ids.issubset(human_by_pair):
        raise ValueError(
            "Formal IAA is not a unique 12-paper subset of the main 48-paper sample."
        )
    if set(human_by_pair) != set(llm_by_pair):
        raise ValueError("Main human and LLM pair_id sets do not match.")

    paper_dir.mkdir(parents=True, exist_ok=True)
    records: list[dict[str, Any]] = []
    for number, pair_id in enumerate(sorted(iaa_pair_ids), start=1):
        human = human_by_pair[pair_id]
        llm = llm_by_pair[pair_id]
        if not as_bool(llm["schema_valid"]) or as_bool(llm["decision_language_detected"]):
            raise ValueError(
                f"IAA LLM review failed schema or decision-language validation: {pair_id}"
            )

        paper_code = f"P{number:03d}"
        source_pdf = resolve_path(human["paper_pdf_path"])
        output_pdf = paper_dir / f"{paper_code}.pdf"
        shutil.copyfile(source_pdf, output_pdf)
        records.append(
            {
                "paper_code": paper_code,
                "title": human["title"],
                "dataset": human["dataset"],
                "score_tier": human["score_tier"],
                "human_review": normalise_section_headings(human["review_text"]),
                "llm_review": normalise_section_headings(llm["generated_review_text"]),
                "source_pdf": source_pdf,
                "output_pdf": output_pdf,
                "generation_model": llm["model_returned"],
                "generation_prompt": llm["prompt_version"],
            }
        )
    return records


def thin_border() -> Border:
    side = Side(style="thin", color=COLOURS["border"])
    return Border(left=side, right=side, top=side, bottom=side)


def style_title(cell: Any, fill_colour: str) -> None:
    cell.fill = PatternFill("solid", fgColor=fill_colour)
    cell.font = Font(color="FFFFFF", bold=True, size=16)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def add_instructions_sheet(workbook: Workbook, paper_codes: list[str]) -> None:
    sheet = workbook.active
    sheet.title = "Instructions"
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:H2")
    sheet["A1"] = "Formal IAA Annotation Workbook"
    style_title(sheet["A1"], COLOURS["primary"])
    sheet.row_dimensions[1].height = 26
    sheet.row_dimensions[2].height = 12

    sheet["A4"] = "Annotator ID"
    sheet["A4"].font = Font(bold=True, color=COLOURS["text"])
    sheet["B4"] = ""
    sheet["B4"].fill = PatternFill("solid", fgColor=COLOURS["input"])
    sheet["B4"].border = thin_border()
    sheet["B4"].alignment = Alignment(horizontal="center")

    instructions = [
        "Complete all 12 paper sheets independently using the same checklist.",
        "For each paper, read the PDF and both displayed reviews. Review 1 is the Human review; Review 2 is the LLM review.",
        "Score each review on its own evidence; do not assign a label merely because the other review is stronger or weaker.",
        "Use exactly: yes, partial, no, or not_applicable. Answer cells provide a dropdown list.",
        "Answer cells are required. Evidence/notes are optional, but can briefly explain a difficult, partial, or not_applicable judgement.",
        "Do not make accept/reject decisions. Save your copy as formal_iaa_<annotator_id>.xlsx.",
    ]
    sheet["A6"] = "Workflow"
    sheet["A6"].font = Font(bold=True, size=12, color=COLOURS["primary"])
    for row_index, instruction in enumerate(instructions, start=7):
        sheet[f"A{row_index}"] = f"{row_index - 6}."
        sheet[f"B{row_index}"] = instruction
        sheet.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=8)
        sheet[f"B{row_index}"].alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row_index].height = 28

    sheet["A15"] = "Answer"
    sheet["B15"] = "Meaning"
    for cell in sheet[15][:2]:
        cell.fill = PatternFill("solid", fgColor=COLOURS["header"])
        cell.font = Font(bold=True, color=COLOURS["text"])
        cell.border = thin_border()
    answer_meanings = [
        ("yes", "Clearly and specifically satisfies the item."),
        ("partial", "Addresses the item incompletely, vaguely, or with limited evidence."),
        ("no", "The item is relevant, but the review does not address it adequately."),
        ("not_applicable", "The item is genuinely irrelevant to this paper or review context."),
    ]
    for row_index, (answer, meaning) in enumerate(answer_meanings, start=16):
        sheet[f"A{row_index}"] = answer
        sheet[f"B{row_index}"] = meaning
        sheet.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=8)
        for cell in sheet[row_index][:8]:
            cell.border = thin_border()

    sheet["A22"] = "Paper sheets"
    sheet["A22"].font = Font(bold=True, size=12, color=COLOURS["primary"])
    for offset, paper_code in enumerate(paper_codes):
        row = 23 + offset // 4
        column = 1 + (offset % 4) * 2
        cell = sheet.cell(row=row, column=column, value=paper_code)
        cell.hyperlink = f"#{paper_code}!A1"
        cell.style = "Hyperlink"

    sheet["A28"] = "Expected judgements"
    sheet["B28"] = 600
    sheet["A29"] = "Completion check"
    sheet["B29"] = "All 12 sheets × 2 reviews × 25 items"
    for row in range(28, 30):
        sheet[f"A{row}"].font = Font(bold=True, color=COLOURS["text"])
        sheet[f"A{row}"].fill = PatternFill("solid", fgColor=COLOURS["primary_light"])
        sheet[f"B{row}"].fill = PatternFill("solid", fgColor=COLOURS["input"])
        sheet[f"A{row}"].border = sheet[f"B{row}"].border = thin_border()

    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 24
    for column in "CDEFGH":
        sheet.column_dimensions[column].width = 15
    sheet.freeze_panes = "A6"


def add_guideline_sheet(
    workbook: Workbook,
    checklist_rows: list[dict[str, str]],
    guideline_text: str,
) -> None:
    sheet = workbook.create_sheet("Guideline")
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:J2")
    sheet["A1"] = "Checklist Guideline v0.3"
    style_title(sheet["A1"], COLOURS["primary"])

    headers = [
        "Question ID",
        "Dimension",
        "Question",
        "Context required",
        "Definition",
        "Yes",
        "Partial",
        "No",
        "Not applicable",
        "Boundary note",
    ]
    header_row = 4
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column, value=header)
        cell.fill = PatternFill("solid", fgColor=COLOURS["header"])
        cell.font = Font(bold=True, color=COLOURS["text"])
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = thin_border()

    for row_index, item in enumerate(checklist_rows, start=header_row + 1):
        parsed = parse_item_guideline(guideline_text, item["question_id"])
        values = [
            item["question_id"],
            item["dimension"],
            parsed["Question"] or item["candidate_question"],
            parsed["Context required"],
            parsed["Definition"],
            parsed["Yes"],
            parsed["Partial"],
            parsed["No"],
            parsed["Not applicable"],
            parsed["Boundary note"],
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin_border()
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=COLOURS["stripe"])
        sheet.row_dimensions[row_index].height = 96

    boundary_start = header_row + len(checklist_rows) + 3
    sheet.cell(row=boundary_start, column=1, value="Pairwise boundary rules").font = Font(
        bold=True, size=12, color=COLOURS["primary"]
    )
    boundary_headers = ["Item pair", "Main distinction", "When evidence supports both"]
    for column, header in enumerate(boundary_headers, start=1):
        cell = sheet.cell(row=boundary_start + 1, column=column, value=header)
        cell.fill = PatternFill("solid", fgColor=COLOURS["header"])
        cell.font = Font(bold=True, color=COLOURS["text"])
        cell.border = thin_border()
    for row_offset, boundary in enumerate(parse_pairwise_boundaries(guideline_text), start=2):
        for column, value in enumerate(boundary, start=1):
            cell = sheet.cell(row=boundary_start + row_offset, column=column, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin_border()
        sheet.row_dimensions[boundary_start + row_offset].height = 60

    widths = [14, 18, 55, 28, 45, 45, 45, 45, 45, 55]
    for column_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + column_index)].width = width
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:J{header_row + len(checklist_rows)}"


def add_answer_validation(sheet: Any, cell_range: str) -> None:
    validation = DataValidation(
        type="list",
        formula1='"yes,partial,no,not_applicable"',
        allow_blank=True,
    )
    validation.error = "Choose yes, partial, no, or not_applicable."
    validation.errorTitle = "Invalid answer"
    validation.prompt = "Select one checklist label."
    validation.promptTitle = "Checklist answer"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    sheet.add_data_validation(validation)
    validation.add(cell_range)


def add_answer_conditional_formatting(sheet: Any, cell_range: str) -> None:
    for answer, colour in [
        ("yes", COLOURS["yes"]),
        ("partial", COLOURS["partial"]),
        ("no", COLOURS["no"]),
        ("not_applicable", COLOURS["na"]),
    ]:
        sheet.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="equal",
                formula=[f'"{answer}"'],
                fill=PatternFill("solid", fgColor=colour),
            ),
        )


def add_paper_sheet(
    workbook: Workbook,
    paper: dict[str, Any],
    checklist_rows: list[dict[str, str]],
) -> None:
    sheet = workbook.create_sheet(paper["paper_code"])
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:G1")
    sheet["A1"] = f"{paper['paper_code']} | {paper['title']}"
    style_title(sheet["A1"], COLOURS["primary"])
    sheet.row_dimensions[1].height = 28

    sheet["A2"] = "Paper PDF"
    sheet["A2"].font = Font(bold=True, color=COLOURS["text"])
    sheet["B2"] = f"Open {paper['paper_code']}.pdf"
    sheet["B2"].hyperlink = f"papers/{paper['paper_code']}.pdf"
    sheet["B2"].style = "Hyperlink"

    sheet.merge_cells("A4:C4")
    sheet.merge_cells("E4:G4")
    sheet["A4"] = "Review 1"
    sheet["E4"] = "Review 2"
    for coordinate, colour in [("A4", COLOURS["human"]), ("E4", COLOURS["llm"])]:
        cell = sheet[coordinate]
        cell.fill = PatternFill("solid", fgColor=colour)
        cell.font = Font(bold=True, size=12, color=COLOURS["text"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()

    sheet.merge_cells("A5:C21")
    sheet.merge_cells("E5:G21")
    sheet["A5"] = paper["human_review"]
    sheet["E5"] = paper["llm_review"]
    for coordinate in ("A5", "E5"):
        cell = sheet[coordinate]
        cell.font = Font(size=9, color=COLOURS["text"])
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = thin_border()
    for row in range(5, 22):
        sheet.row_dimensions[row].height = 22

    sheet.merge_cells("A22:G22")
    sheet["A22"] = "Apply every item independently to each review; use the PDF whenever paper context is required."
    sheet["A22"].font = Font(italic=True, color=COLOURS["muted"])
    sheet["A22"].alignment = Alignment(horizontal="center")

    headers = [
        "Item",
        "Dimension",
        "Question",
        "Review 1 answer",
        "Review 1 optional evidence / note",
        "Review 2 answer",
        "Review 2 optional evidence / note",
    ]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=TABLE_HEADER_ROW, column=column, value=header)
        cell.fill = PatternFill("solid", fgColor=COLOURS["header"])
        cell.font = Font(bold=True, color=COLOURS["text"])
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border = thin_border()
    sheet["D24"].fill = PatternFill("solid", fgColor=COLOURS["human"])
    sheet["F24"].fill = PatternFill("solid", fgColor=COLOURS["llm"])

    for row_index, item in enumerate(checklist_rows, start=FIRST_QUESTION_ROW):
        values = [item["question_id"], item["dimension"], item["candidate_question"]]
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column, value=value)
        for column in range(1, 8):
            cell = sheet.cell(row=row_index, column=column)
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top",
                horizontal="center" if column in {1, 2, 4, 6} else "left",
            )
            cell.border = thin_border()
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=COLOURS["stripe"])
        for column in (4, 6):
            sheet.cell(row=row_index, column=column).fill = PatternFill(
                "solid", fgColor=COLOURS["input"]
            )
        estimated_lines = max(2, math.ceil(len(item["candidate_question"]) / 72))
        sheet.row_dimensions[row_index].height = min(84, max(42, estimated_lines * 18))

    add_answer_validation(sheet, f"D{FIRST_QUESTION_ROW}:D{LAST_QUESTION_ROW}")
    add_answer_validation(sheet, f"F{FIRST_QUESTION_ROW}:F{LAST_QUESTION_ROW}")
    add_answer_conditional_formatting(sheet, f"D{FIRST_QUESTION_ROW}:D{LAST_QUESTION_ROW}")
    add_answer_conditional_formatting(sheet, f"F{FIRST_QUESTION_ROW}:F{LAST_QUESTION_ROW}")

    widths = [11, 18, 85, 20, 44, 20, 44]
    for column_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + column_index)].width = width
    sheet.freeze_panes = f"A{FIRST_QUESTION_ROW}"
    sheet.auto_filter.ref = f"A{TABLE_HEADER_ROW}:G{LAST_QUESTION_ROW}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.tabColor = COLOURS["primary"]


def build_workbook(
    paper_records: list[dict[str, Any]],
    checklist_rows: list[dict[str, str]],
    guideline_text: str,
    output_path: Path,
) -> None:
    workbook = Workbook()
    add_instructions_sheet(workbook, [paper["paper_code"] for paper in paper_records])
    add_guideline_sheet(workbook, checklist_rows, guideline_text)
    for paper in paper_records:
        add_paper_sheet(workbook, paper, checklist_rows)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def build_annotator_guideline(guideline_text: str, output_path: Path) -> None:
    preface = """# Formal IAA Guideline v0.1

## Workbook layout

- `Review 1` is the Human review.
- `Review 2` is the LLM review.
- Every answer cell is required and must be `yes`, `partial`, `no`, or
  `not_applicable`.
- Evidence/notes are optional. Use them when a judgement is difficult, partial,
  not applicable, or would benefit from a short explanation.
- Score Review 1 and Review 2 independently against the same checklist and
  paper. Do not make accept/reject decisions.

---

"""
    output_path.write_text(preface + guideline_text, encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build a formal IAA annotation XLSX.")
    parser.add_argument("--human-path", default=DEFAULT_HUMAN_PATH)
    parser.add_argument("--iaa-path", default=DEFAULT_IAA_PATH)
    parser.add_argument("--llm-path", default=DEFAULT_LLM_PATH)
    parser.add_argument("--checklist-path", default=DEFAULT_CHECKLIST_PATH)
    parser.add_argument("--guideline-path", default=DEFAULT_GUIDELINE_PATH)
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--workbook-name", default=DEFAULT_WORKBOOK_NAME)
    parser.add_argument(
        "--annotator-guideline-name",
        default=DEFAULT_ANNOTATOR_GUIDELINE_NAME,
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    human_rows = read_csv_rows(resolve_path(args.human_path))
    iaa_rows = read_csv_rows(resolve_path(args.iaa_path))
    llm_rows = read_csv_rows(resolve_path(args.llm_path))
    checklist_rows = read_csv_rows(resolve_path(args.checklist_path))
    guideline_text = resolve_path(args.guideline_path).read_text(encoding="utf-8")
    output_dir = resolve_path(args.output_dir)
    paper_dir = output_dir / "papers"

    if len(human_rows) != 48 or len(llm_rows) != 48 or len(iaa_rows) != 12:
        raise ValueError(
            "Formal IAA expects 48 main human reviews, 48 main LLM reviews, "
            "and 12 IAA papers."
        )
    if len(checklist_rows) != 25:
        raise ValueError(f"Checklist expected 25 items; found {len(checklist_rows)}.")

    paper_records = build_paper_records(human_rows, iaa_rows, llm_rows, paper_dir)
    workbook_path = output_dir / args.workbook_name
    annotator_guideline_path = output_dir / args.annotator_guideline_name
    build_workbook(paper_records, checklist_rows, guideline_text, workbook_path)
    build_annotator_guideline(guideline_text, annotator_guideline_path)

    dataset_tier_counts = Counter(
        (paper["dataset"], paper["score_tier"]) for paper in paper_records
    )
    pdf_hashes_match = all(
        sha256(paper["source_pdf"]) == sha256(paper["output_pdf"])
        for paper in paper_records
    )
    if set(dataset_tier_counts.values()) != {2} or len(dataset_tier_counts) != 6:
        raise ValueError(
            "Formal IAA is unbalanced across dataset and tier: "
            f"{dict(dataset_tier_counts)}"
        )
    if not pdf_hashes_match:
        raise ValueError("Formal IAA PDF hash validation failed.")

    summary = {
        "papers": 12,
        "human_reviews": 12,
        "llm_reviews": 12,
        "checklist_items_per_review": 25,
        "expected_judgements_per_annotator": 600,
        "workbook_sheets": ["Instructions", "Guideline"]
        + [paper["paper_code"] for paper in paper_records],
        "paper_order": [paper["paper_code"] for paper in paper_records],
        "first_review": "human",
        "second_review": "llm",
        "dataset_tier_paper_counts": {
            f"{dataset}::{tier}": count
            for (dataset, tier), count in sorted(dataset_tier_counts.items())
        },
        "pdf_hashes_match_source": pdf_hashes_match,
        "generation_models": sorted({paper["generation_model"] for paper in paper_records}),
        "generation_prompts": sorted({paper["generation_prompt"] for paper in paper_records}),
        "workbook_path": str(workbook_path.relative_to(PROJECT_ROOT)),
        "annotator_guideline_path": str(
            annotator_guideline_path.relative_to(PROJECT_ROOT)
        ),
        "paper_dir": str(paper_dir.relative_to(PROJECT_ROOT)),
    }
    summary_path = output_dir / "formal_iaa_v0_1_summary.json"
    summary_path.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
