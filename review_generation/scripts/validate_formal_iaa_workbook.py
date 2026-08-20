"""Validate whether a returned formal IAA XLSX is ready for agreement analysis."""

from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_CHECKLIST_PATH = "artifacts/checklist.csv"
VALID_ANSWERS = {"yes", "partial", "no", "not_applicable"}
PAPER_SHEETS = [f"P{index:03d}" for index in range(1, 13)]
FIRST_QUESTION_ROW = 25
LAST_QUESTION_ROW = 49


def resolve_path(path_value: str | Path) -> Path:
    path = Path(path_value)
    return path if path.is_absolute() else PROJECT_ROOT / path


def display_path(path: Path) -> str:
    try:
        return str(path.relative_to(PROJECT_ROOT))
    except ValueError:
        return str(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Validate a formal IAA XLSX workbook.")
    parser.add_argument("workbook_path")
    parser.add_argument(
        "--allow-incomplete",
        action="store_true",
        help="Allow blank answers when validating the master workbook before distribution.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workbook_path = resolve_path(args.workbook_path)
    workbook = load_workbook(workbook_path, data_only=False)
    expected_sheets = ["Instructions", "Guideline", *PAPER_SHEETS]
    sheet_names_match = workbook.sheetnames == expected_sheets
    annotator_id = str(workbook["Instructions"]["B4"].value or "").strip()

    answers: list[str] = []
    filled_optional_note_n = 0
    question_ids_by_sheet: dict[str, list[str]] = {}
    validation_missing_sheets: list[str] = []
    review_label_mismatch_sheets: list[str] = []
    for sheet_name in PAPER_SHEETS:
        sheet = workbook[sheet_name]
        if sheet["A4"].value != "Review 1" or sheet["E4"].value != "Review 2":
            review_label_mismatch_sheets.append(sheet_name)
        question_ids_by_sheet[sheet_name] = [
            str(sheet.cell(row=row, column=1).value or "")
            for row in range(FIRST_QUESTION_ROW, LAST_QUESTION_ROW + 1)
        ]
        answer_coordinates = {
            f"{column}{row}"
            for column in ("D", "F")
            for row in range(FIRST_QUESTION_ROW, LAST_QUESTION_ROW + 1)
        }
        validated_coordinates = {
            coordinate
            for coordinate in answer_coordinates
            if any(
                coordinate in validation
                for validation in sheet.data_validations.dataValidation
            )
        }
        if validated_coordinates != answer_coordinates:
            validation_missing_sheets.append(sheet_name)
        for row in range(FIRST_QUESTION_ROW, LAST_QUESTION_ROW + 1):
            for answer_column, note_column in [(4, 5), (6, 7)]:
                answer = str(sheet.cell(row=row, column=answer_column).value or "").strip().lower()
                note = str(sheet.cell(row=row, column=note_column).value or "").strip()
                answers.append(answer)
                if note:
                    filled_optional_note_n += 1

    invalid_answers = sorted(
        {answer for answer in answers if answer and answer not in VALID_ANSWERS}
    )
    missing_answer_n = sum(not answer for answer in answers)
    reference_ids = question_ids_by_sheet[PAPER_SHEETS[0]]
    question_ids_consistent = all(
        question_ids == reference_ids for question_ids in question_ids_by_sheet.values()
    ) and len(set(reference_ids)) == 25

    structurally_valid = (
        sheet_names_match
        and len(answers) == 600
        and not invalid_answers
        and question_ids_consistent
        and not validation_missing_sheets
        and not review_label_mismatch_sheets
    )
    complete = structurally_valid and bool(annotator_id) and missing_answer_n == 0
    report = {
        "workbook_path": display_path(workbook_path),
        "structurally_valid": structurally_valid,
        "complete": complete,
        "annotator_id": annotator_id,
        "sheet_names_match": sheet_names_match,
        "judgement_cells": len(answers),
        "answer_counts": dict(Counter(answer or "<blank>" for answer in answers)),
        "missing_answer_n": missing_answer_n,
        "invalid_answers": invalid_answers,
        "filled_optional_note_n": filled_optional_note_n,
        "question_ids_consistent": question_ids_consistent,
        "validation_missing_sheets": validation_missing_sheets,
        "review_label_mismatch_sheets": review_label_mismatch_sheets,
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))

    if not structurally_valid:
        raise SystemExit("Formal IAA workbook failed structural validation.")
    if not args.allow_incomplete and not complete:
        raise SystemExit("Formal IAA workbook has not been completed.")


if __name__ == "__main__":
    main()
