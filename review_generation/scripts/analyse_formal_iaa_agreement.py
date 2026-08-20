"""Calculate Cohen's kappa for two annotators' formal IAA workbooks."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from statsmodels.stats.inter_rater import cohens_kappa as statsmodels_cohens_kappa


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_ANNOTATOR_A_PATH = "inputs/human_iaa/annotator_a.xlsx"
DEFAULT_ANNOTATOR_B_PATH = "inputs/human_iaa/annotator_b.xlsx"
DEFAULT_OUTPUT_DIR = "outputs/human_iaa/analysis"

PAPER_SHEETS = [f"P{index:03d}" for index in range(1, 13)]
FIRST_QUESTION_ROW = 25
LAST_QUESTION_ROW = 49
NOMINAL_LABELS = ["yes", "partial", "no", "not_applicable"]
ORDINAL_LABELS = ["no", "partial", "yes"]
VALID_LABELS = set(NOMINAL_LABELS)
DIMENSION_ORDER = [
    "Coverage",
    "Substance",
    "Reasoning",
    "Grounding",
    "Constructiveness",
    "Independence",
    "Specificity",
    "Clarity",
    "Ethics",
]
REVIEW_COLUMNS = {
    "human": {"review_number": 1, "answer_column": 4, "note_column": 5},
    "llm": {"review_number": 2, "answer_column": 6, "note_column": 7},
}


def resolve_path(path_value: str | Path) -> Path:
    path = Path(path_value)
    return path if path.is_absolute() else PROJECT_ROOT / path


def display_path(path: Path) -> str:
    try:
        return str(path.relative_to(PROJECT_ROOT))
    except ValueError:
        return str(path)


def cell_text(value: Any) -> str:
    return str(value or "").strip()


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for block in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def parse_paper_title(heading: str, paper_code: str) -> str:
    prefix = f"{paper_code} | "
    if not heading.startswith(prefix):
        raise ValueError(
            f"{paper_code} has an invalid A1 heading: expected it to start with "
            f"{prefix!r}, got {heading!r}."
        )
    title = heading[len(prefix) :].strip()
    if not title:
        raise ValueError(f"{paper_code} has an empty paper title.")
    return title


def read_detailed_judgements(workbook: Any) -> dict[tuple[str, str, str], str]:
    if "Detailed Judgements" not in workbook.sheetnames:
        return {}

    sheet = workbook["Detailed Judgements"]
    headers = [cell_text(cell.value) for cell in sheet[1]]
    required = ["Paper ID", "Reviewer", "Question ID", "Review Result"]
    if any(header not in headers for header in required):
        raise ValueError("Detailed Judgements is missing required columns.")
    positions = {header: headers.index(header) for header in required}
    records: dict[tuple[str, str, str], str] = {}
    for values in sheet.iter_rows(min_row=2, values_only=True):
        paper_code = cell_text(values[positions["Paper ID"]])
        reviewer = cell_text(values[positions["Reviewer"]]).lower()
        question_id = cell_text(values[positions["Question ID"]])
        answer = cell_text(values[positions["Review Result"]]).lower()
        if not any([paper_code, reviewer, question_id, answer]):
            continue
        key = (paper_code, reviewer, question_id)
        if key in records:
            raise ValueError(f"Detailed Judgements contains a duplicate key: {key}")
        records[key] = answer
    return records


def read_workbook_annotations(
    path: Path,
    *,
    check_detailed_judgements: bool = True,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    missing_sheets = [name for name in PAPER_SHEETS if name not in workbook.sheetnames]
    if missing_sheets:
        raise ValueError(f"{path.name} is missing paper sheets: {missing_sheets}")

    annotator_id = cell_text(workbook["Instructions"]["B4"].value)
    if not annotator_id:
        raise ValueError(f"{path.name} is missing an annotator ID.")

    detailed_judgements_present = "Detailed Judgements" in workbook.sheetnames
    detailed = (
        read_detailed_judgements(workbook)
        if check_detailed_judgements
        else {}
    )
    rows: list[dict[str, Any]] = []
    detailed_mismatches: list[str] = []
    for paper_code in PAPER_SHEETS:
        sheet = workbook[paper_code]
        paper_title = parse_paper_title(cell_text(sheet["A1"].value), paper_code)
        if cell_text(sheet["A4"].value) != "Review 1":
            raise ValueError(f"{path.name} {paper_code} has an invalid Review 1 label.")
        if cell_text(sheet["E4"].value) != "Review 2":
            raise ValueError(f"{path.name} {paper_code} has an invalid Review 2 label.")

        for row_number in range(FIRST_QUESTION_ROW, LAST_QUESTION_ROW + 1):
            question_id = cell_text(sheet.cell(row=row_number, column=1).value)
            dimension = cell_text(sheet.cell(row=row_number, column=2).value)
            question = cell_text(sheet.cell(row=row_number, column=3).value)
            if not question_id or not dimension or not question:
                raise ValueError(
                    f"{path.name} {paper_code} row {row_number} is missing question "
                    "metadata."
                )

            for review_source, columns in REVIEW_COLUMNS.items():
                answer = cell_text(
                    sheet.cell(row=row_number, column=columns["answer_column"]).value
                ).lower()
                note = cell_text(
                    sheet.cell(row=row_number, column=columns["note_column"]).value
                )
                if answer not in VALID_LABELS:
                    raise ValueError(
                        f"{path.name} {paper_code} {review_source} {question_id}"
                        f" contains an invalid or empty answer: {answer!r}."
                    )
                detailed_key = (paper_code, review_source, question_id)
                if detailed and detailed.get(detailed_key) != answer:
                    detailed_mismatches.append(
                        f"{detailed_key}: sheet={answer}, detailed={detailed.get(detailed_key)!r}"
                    )
                rows.append(
                    {
                        "paper_code": paper_code,
                        "paper_title": paper_title,
                        "review_source": review_source,
                        "review_number": columns["review_number"],
                        "question_id": question_id,
                        "dimension": dimension,
                        "question": question,
                        "answer": answer,
                        "note": note,
                    }
                )

    if len(rows) != 600:
        raise ValueError(f"{path.name} expected 600 judgements; found {len(rows)}.")
    if detailed and len(detailed) != 600:
        raise ValueError(
            f"{path.name} Detailed Judgements expected 600 rows; "
            f"found {len(detailed)}."
        )
    if detailed_mismatches:
        raise ValueError(
            f"{path.name} paper sheets do not match Detailed Judgements: "
            + "; ".join(detailed_mismatches[:10])
        )

    metadata = {
        "path": display_path(path),
        "file_name": path.name,
        "annotator_id": annotator_id,
        "sha256": file_sha256(path),
        "sheet_names": workbook.sheetnames,
        "extra_sheets": [
            name
            for name in workbook.sheetnames
            if name not in {"Instructions", "Guideline", *PAPER_SHEETS}
        ],
        "judgements": len(rows),
        "answer_counts": dict(Counter(row["answer"] for row in rows)),
        "filled_notes": sum(bool(row["note"]) for row in rows),
        "annotation_source": "P001-P012 paper sheets",
        "detailed_judgements_present": detailed_judgements_present,
        "detailed_judgements_checked": bool(
            detailed_judgements_present and check_detailed_judgements
        ),
        "detailed_judgements_ignored": bool(
            detailed_judgements_present and not check_detailed_judgements
        ),
        "detailed_judgements_match": not detailed_mismatches if detailed else None,
    }
    return rows, metadata


def align_annotations(
    annotator_a_rows: list[dict[str, Any]],
    annotator_b_rows: list[dict[str, Any]],
    annotator_a_id: str,
    annotator_b_id: str,
) -> list[dict[str, Any]]:
    key_fields = ("paper_code", "review_source", "question_id")
    by_a = {tuple(row[field] for field in key_fields): row for row in annotator_a_rows}
    by_b = {tuple(row[field] for field in key_fields): row for row in annotator_b_rows}
    if len(by_a) != 600 or len(by_b) != 600:
        raise ValueError("At least one workbook contains duplicate annotation keys.")
    if set(by_a) != set(by_b):
        missing_in_b = sorted(set(by_a) - set(by_b))
        missing_in_a = sorted(set(by_b) - set(by_a))
        raise ValueError(
            "The two workbooks cannot be aligned one-to-one; "
            f"B is missing {missing_in_b[:5]}, and A is missing {missing_in_a[:5]}."
        )

    aligned: list[dict[str, Any]] = []
    for key in sorted(by_a):
        left = by_a[key]
        right = by_b[key]
        for field in ("paper_title", "review_number", "dimension", "question"):
            if left[field] != right[field]:
                raise ValueError(
                    f"{key}: {field} differs between the two workbooks."
                )
        answer_a = left["answer"]
        answer_b = right["answer"]
        aligned.append(
            {
                "paper_code": left["paper_code"],
                "paper_title": left["paper_title"],
                "review_source": left["review_source"],
                "review_number": left["review_number"],
                "question_id": left["question_id"],
                "dimension": left["dimension"],
                "question": left["question"],
                "annotator_a_id": annotator_a_id,
                "annotator_a_answer": answer_a,
                "annotator_a_note": left["note"],
                "annotator_b_id": annotator_b_id,
                "annotator_b_answer": answer_b,
                "annotator_b_note": right["note"],
                "exact_agreement": answer_a == answer_b,
                "jointly_applicable": (
                    answer_a != "not_applicable" and answer_b != "not_applicable"
                ),
                "disagreement_pair": (
                    "agree"
                    if answer_a == answer_b
                    else " ↔ ".join(sorted([answer_a, answer_b]))
                ),
            }
        )
    return aligned


def unweighted_kappa(
    pairs: list[tuple[str, str]], labels: list[str]
) -> tuple[float | None, float, float]:
    if not pairs:
        return None, 0.0, 0.0
    total = len(pairs)
    counts_a = Counter(left for left, _ in pairs)
    counts_b = Counter(right for _, right in pairs)
    observed = sum(left == right for left, right in pairs) / total
    expected = sum(
        (counts_a[label] / total) * (counts_b[label] / total) for label in labels
    )
    if expected == 1.0:
        return None, observed, expected
    return (observed - expected) / (1.0 - expected), observed, expected


def weighted_kappa(
    pairs: list[tuple[str, str]], labels: list[str], weighting: str
) -> float | None:
    if not pairs:
        return None
    if weighting not in {"linear", "quadratic"}:
        raise ValueError(f"Unsupported weighting: {weighting}")

    total = len(pairs)
    last_index = len(labels) - 1
    positions = {label: index for index, label in enumerate(labels)}
    counts_a = Counter(left for left, _ in pairs)
    counts_b = Counter(right for _, right in pairs)

    def weight(left: str, right: str) -> float:
        distance = abs(positions[left] - positions[right]) / last_index
        penalty = distance if weighting == "linear" else distance**2
        return 1.0 - penalty

    observed = sum(weight(left, right) for left, right in pairs) / total
    expected = sum(
        (counts_a[left] / total)
        * (counts_b[right] / total)
        * weight(left, right)
        for left in labels
        for right in labels
    )
    if expected == 1.0:
        return None
    return (observed - expected) / (1.0 - expected)


def rounded(value: float | None) -> float | str:
    return "" if value is None else round(value, 6)


def calculate_scope_stats(
    scope_type: str,
    scope_value: str,
    rows: list[dict[str, Any]],
) -> dict[str, Any]:
    all_pairs = [
        (row["annotator_a_answer"], row["annotator_b_answer"]) for row in rows
    ]
    applicable_pairs = [
        pair
        for pair in all_pairs
        if pair[0] != "not_applicable" and pair[1] != "not_applicable"
    ]
    nominal_kappa, observed, expected = unweighted_kappa(all_pairs, NOMINAL_LABELS)
    applicable_kappa, applicable_observed, applicable_expected = unweighted_kappa(
        applicable_pairs, ORDINAL_LABELS
    )
    return {
        "scope_type": scope_type,
        "scope_value": scope_value,
        "judgements_n": len(all_pairs),
        "exact_agreement_n": sum(left == right for left, right in all_pairs),
        "exact_agreement_percent": round(observed * 100, 2),
        "expected_agreement_percent": round(expected * 100, 2),
        "cohen_kappa_4_labels": rounded(nominal_kappa),
        "jointly_applicable_n": len(applicable_pairs),
        "jointly_applicable_agreement_percent": round(applicable_observed * 100, 2),
        "jointly_applicable_expected_percent": round(applicable_expected * 100, 2),
        "cohen_kappa_3_labels": rounded(applicable_kappa),
        "linear_weighted_kappa_3_labels": rounded(
            weighted_kappa(applicable_pairs, ORDINAL_LABELS, "linear")
        ),
        "quadratic_weighted_kappa_3_labels": rounded(
            weighted_kappa(applicable_pairs, ORDINAL_LABELS, "quadratic")
        ),
        "both_not_applicable_n": sum(
            left == right == "not_applicable" for left, right in all_pairs
        ),
        "one_not_applicable_n": sum(
            (left == "not_applicable") != (right == "not_applicable")
            for left, right in all_pairs
        ),
    }


def add_conventional_kappa_ci(
    overall: dict[str, Any], aligned: list[dict[str, Any]]
) -> None:
    """Add the conventional asymptotic SE and 95% CI.

    Treat the 600 item judgements as the units of analysis.
    """

    counts = Counter(
        (row["annotator_a_answer"], row["annotator_b_answer"]) for row in aligned
    )
    matrix = [
        [counts[(left, right)] for right in NOMINAL_LABELS]
        for left in NOMINAL_LABELS
    ]
    result = statsmodels_cohens_kappa(matrix)
    overall["cohen_kappa_4_labels_asymptotic_se"] = round(
        float(result.std_kappa), 6
    )
    overall["cohen_kappa_4_labels_ci95_lower"] = round(
        float(result.kappa_low), 6
    )
    overall["cohen_kappa_4_labels_ci95_upper"] = round(
        float(result.kappa_upp), 6
    )


def group_stats(
    aligned: list[dict[str, Any]], field: str, scope_type: str
) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in aligned:
        grouped[str(row[field])].append(row)
    return [
        calculate_scope_stats(scope_type, value, rows)
        for value, rows in sorted(grouped.items())
    ]


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def build_confusion_rows(aligned: list[dict[str, Any]]) -> list[dict[str, Any]]:
    counts = Counter(
        (row["annotator_a_answer"], row["annotator_b_answer"]) for row in aligned
    )
    rows: list[dict[str, Any]] = []
    for annotator_a_label in NOMINAL_LABELS:
        row: dict[str, Any] = {"annotator_a_label": annotator_a_label}
        for annotator_b_label in NOMINAL_LABELS:
            row[f"annotator_b_{annotator_b_label}"] = counts[
                (annotator_a_label, annotator_b_label)
            ]
        row["row_total"] = sum(
            counts[(annotator_a_label, label)] for label in NOMINAL_LABELS
        )
        rows.append(row)
    rows.append(
        {
            "annotator_a_label": "column_total",
            **{
                f"annotator_b_{label}": sum(
                    counts[(left, label)] for left in NOMINAL_LABELS
                )
                for label in NOMINAL_LABELS
            },
            "row_total": len(aligned),
        }
    )
    return rows


def build_label_distribution_rows(
    aligned: list[dict[str, Any]], annotator_a_id: str, annotator_b_id: str
) -> list[dict[str, Any]]:
    counts_a = Counter(row["annotator_a_answer"] for row in aligned)
    counts_b = Counter(row["annotator_b_answer"] for row in aligned)
    return [
        {
            "label": label,
            f"{annotator_a_id}_n": counts_a[label],
            f"{annotator_a_id}_percent": round(counts_a[label] / len(aligned) * 100, 2),
            f"{annotator_b_id}_n": counts_b[label],
            f"{annotator_b_id}_percent": round(counts_b[label] / len(aligned) * 100, 2),
        }
        for label in NOMINAL_LABELS
    ]


def top_disagreement_pairs(aligned: list[dict[str, Any]]) -> list[dict[str, Any]]:
    counts = Counter(
        row["disagreement_pair"]
        for row in aligned
        if not row["exact_agreement"]
    )
    total_disagreements = sum(counts.values())
    return [
        {
            "disagreement_pair": pair,
            "n": count,
            "percent_of_disagreements": round(count / total_disagreements * 100, 2),
        }
        for pair, count in counts.most_common()
    ]


def format_metric(value: Any) -> str:
    return "NA" if value == "" or value is None else f"{float(value):.3f}"


def write_readme(
    path: Path,
    overall: dict[str, Any],
    by_source: list[dict[str, Any]],
    by_dimension: list[dict[str, Any]],
    label_distribution: list[dict[str, Any]],
    annotator_a: dict[str, Any],
    annotator_b: dict[str, Any],
    disagreement_pairs: list[dict[str, Any]],
) -> None:
    top_pair = disagreement_pairs[0] if disagreement_pairs else None
    top_pair_text = (
        f"`{top_pair['disagreement_pair']}` ({top_pair['n']} cases)"
        if top_pair
        else "none"
    )
    dimension_by_name = {row["scope_value"]: row for row in by_dimension}
    dimension_lines = [
        "| Dimension | N | Exact agreement | Cohen's κ (4 labels) |",
        "| --- | ---: | ---: | ---: |",
    ]
    for dimension in DIMENSION_ORDER:
        row = dimension_by_name[dimension]
        kappa = (
            "not estimable†"
            if row["cohen_kappa_4_labels"] == ""
            else f"{float(row['cohen_kappa_4_labels']):.3f}"
        )
        dimension_lines.append(
            f"| {dimension} | {row['judgements_n']} | "
            f"{row['exact_agreement_percent']:.2f}% | {kappa} |"
        )
    dimension_table = "\n".join(dimension_lines)

    source_lines = [
        "| Review source | N | Exact agreement | Cohen's κ (4 labels) | Linear weighted κ* |",
        "| --- | ---: | ---: | ---: | ---: |",
    ]
    for row in sorted(by_source, key=lambda item: item["scope_value"]):
        source_label = "LLM" if row["scope_value"] == "llm" else "Human"
        source_lines.append(
            f"| {source_label} | {row['judgements_n']} | "
            f"{row['exact_agreement_percent']:.2f}% | "
            f"{format_metric(row['cohen_kappa_4_labels'])} | "
            f"{format_metric(row['linear_weighted_kappa_3_labels'])} |"
        )
    source_table = "\n".join(source_lines)

    distribution_lines = [
        f"| Label | {annotator_a['annotator_id']} n (%) | "
        f"{annotator_b['annotator_id']} n (%) |",
        "| --- | ---: | ---: |",
    ]
    annotator_a_n = f"{annotator_a['annotator_id']}_n"
    annotator_a_percent = f"{annotator_a['annotator_id']}_percent"
    annotator_b_n = f"{annotator_b['annotator_id']}_n"
    annotator_b_percent = f"{annotator_b['annotator_id']}_percent"
    for row in label_distribution:
        distribution_lines.append(
            f"| `{row['label']}` | "
            f"{row[annotator_a_n]} ({row[annotator_a_percent]:.2f}%) | "
            f"{row[annotator_b_n]} ({row[annotator_b_percent]:.2f}%) |"
        )
    distribution_table = "\n".join(distribution_lines)

    def detailed_status(metadata: dict[str, Any]) -> str:
        if metadata["detailed_judgements_ignored"]:
            return "present but intentionally ignored because the P001–P012 sheets are the updated annotation source"
        if metadata["detailed_judgements_checked"]:
            return "checked and consistent with the P001–P012 sheets"
        return "not present; the P001–P012 sheets were used directly"

    annotator_a_detailed_status = detailed_status(annotator_a)
    annotator_b_detailed_status = detailed_status(annotator_b)

    content = f"""# Formal Human IAA Check Results

## Study design

- Annotators: 2 (`{annotator_a['annotator_id']}` and `{annotator_b['annotator_id']}`).
- Sample: 12 papers and 24 reviews (12 Human and 12 LLM reviews).
- Checklist: 25 items applied to every review.
- Analysis unit: one `review × checklist item` judgement.
- Total paired judgements: 600.
- Labels: `yes`, `partial`, `no`, and `not_applicable`.

## Final result

- Primary statistic: unweighted Cohen's kappa across all four labels
  (`yes`, `partial`, `no`, `not_applicable`) =
  **{format_metric(overall['cohen_kappa_4_labels'])}**.
- Conventional asymptotic 95% CI = **[{overall['cohen_kappa_4_labels_ci95_lower']:.3f}, {overall['cohen_kappa_4_labels_ci95_upper']:.3f}]**.
- Exact agreement = **{overall['exact_agreement_percent']:.2f}%**
  ({overall['exact_agreement_n']}/{overall['judgements_n']}).
- Excluding any judgement where either annotator selected `not_applicable`,
  unweighted Cohen's kappa across `yes/partial/no` =
  **{format_metric(overall['cohen_kappa_3_labels'])}**
  (n = {overall['jointly_applicable_n']}).
- Linear weighted kappa across the jointly applicable ordinal labels =
  **{format_metric(overall['linear_weighted_kappa_3_labels'])}**.
- Quadratic weighted kappa across the jointly applicable ordinal labels =
  **{format_metric(overall['quadratic_weighted_kappa_3_labels'])}**.

The four-label unweighted kappa is the primary result because
`not_applicable` is a distinct nominal category and should not be forced into
the `no < partial < yes` ordering. Weighted kappas are supplementary and are
therefore calculated only for jointly applicable `no/partial/yes` judgements.

## Agreement by dimension

{dimension_table}

† Independence has 100% exact agreement, but kappa is not estimable because
both annotators used only one label, leaving no marginal variance.

## Agreement by review source

{source_table}

*Linear weighted kappa excludes rows where either annotator selected
`not_applicable`. The LLM subset has higher exact agreement but lower
unweighted kappa because both annotators used `yes` very frequently, increasing
chance-expected agreement (a prevalence effect).

## Label distribution

{distribution_table}

## Integrity checks

- Annotator A: `{annotator_a['annotator_id']}` — 600 complete judgements.
- Annotator B: `{annotator_b['annotator_id']}` — 600 complete judgements.
- All 600 `paper × review source × checklist item` keys aligned one-to-one.
- Annotation source: the visible `P001`–`P012` paper sheets in each workbook.
- Annotator A `Detailed Judgements`: {annotator_a_detailed_status}.
- Annotator B `Detailed Judgements`: {annotator_b_detailed_status}.
- Invalid or missing labels: 0.
- Both selected `not_applicable`: {overall['both_not_applicable_n']}.
- Only one selected `not_applicable`: {overall['one_not_applicable_n']}.
- Most frequent disagreement type: {top_pair_text}.

Spreadsheet applications may remove dropdown validations or retain stale
derived summary sheets after manual editing. These formatting or derived-sheet
differences do not affect this analysis because it validates and aligns the
600 legal answer values directly from the visible paper sheets.

## Output files

- `overall_summary.csv`: primary and supplementary overall agreement metrics.
- `agreement_by_review_source.csv`: Human versus LLM review agreement.
- `agreement_by_dimension.csv`: agreement for each checklist dimension.
- `agreement_by_item.csv`: agreement for each checklist item (n=24 each;
  interpret cautiously because kappa is unstable with small or highly
  imbalanced groups).
- `agreement_by_paper.csv`: descriptive agreement for each paper.
- `confusion_matrix.csv`: four-label cross-tabulation.
- `label_distribution.csv`: each annotator's label prevalence.
- `disagreement_types.csv`: frequency of label-pair disagreements.
- `disagreements.csv`: all individual disagreements with both evidence notes.
- `aligned_judgements.csv`: complete aligned 600-row audit table.
- `summary.json`: machine-readable results and provenance.

Kappa is prevalence-sensitive, so report exact agreement alongside it rather
than interpreting kappa alone.

## Recommended research reporting

### Main text

Report the study design, exact agreement, the four-label unweighted kappa and
its 95% CI, and linear weighted kappa as a supplementary ordinal result. Include
the dimension-level table or summarise its strongest and weakest dimensions.

Suggested results sentence:

> Across {overall['judgements_n']} paired item-level judgements, exact agreement was
> {overall['exact_agreement_percent']:.1f}%, with an unweighted Cohen's κ of
> {float(overall['cohen_kappa_4_labels']):.2f} (95% CI
> [{overall['cohen_kappa_4_labels_ci95_lower']:.2f}, {overall['cohen_kappa_4_labels_ci95_upper']:.2f}]). Agreement was stronger
> when applicable ratings were treated as ordinal (linear weighted κ =
> {float(overall['linear_weighted_kappa_3_labels']):.2f}),
> and most disagreements occurred between adjacent categories.

Describe the result as usable but heterogeneous operational reliability, not
as uniformly high reliability or a fully validated instrument. Grounding,
Substance, and Constructiveness were less stable than Specificity, Ethics,
Reasoning, and Coverage.

### Appendix or supplementary material

Place the complete 25-item table, confusion matrix, label distributions,
review-source stratification, and disagreement audit in the appendix. Do not
average the 25 item-level kappas. Item-level estimates each use only 24
judgements and are sensitive to prevalence; for example, `COV_1` has 95.83%
exact agreement but κ = 0 because almost all ratings fall in one category.

Human–human IAA assesses whether the guideline can be applied consistently. If
the research also claims that the automated evaluator is reliable, report
a separate automated-evaluator-versus-human agreement analysis on the same 24
reviews; the human–human kappa alone does not validate the automated evaluator.
"""
    path.write_text(content, encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Calculate Cohen's kappa for two formal IAA workbooks."
    )
    parser.add_argument("--annotator-a-path", default=DEFAULT_ANNOTATOR_A_PATH)
    parser.add_argument("--annotator-b-path", default=DEFAULT_ANNOTATOR_B_PATH)
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR)
    parser.add_argument(
        "--ignore-detailed-judgements",
        action="store_true",
        help=(
            "Use the visible P001-P012 paper sheets as the annotation source "
            "without checking the potentially stale derived Detailed Judgements sheet."
        ),
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    annotator_a_path = resolve_path(args.annotator_a_path)
    annotator_b_path = resolve_path(args.annotator_b_path)
    output_dir = resolve_path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    annotator_a_rows, annotator_a_metadata = read_workbook_annotations(
        annotator_a_path,
        check_detailed_judgements=not args.ignore_detailed_judgements,
    )
    annotator_b_rows, annotator_b_metadata = read_workbook_annotations(
        annotator_b_path,
        check_detailed_judgements=not args.ignore_detailed_judgements,
    )
    aligned = align_annotations(
        annotator_a_rows,
        annotator_b_rows,
        annotator_a_metadata["annotator_id"],
        annotator_b_metadata["annotator_id"],
    )

    overall = calculate_scope_stats("overall", "all_600_judgements", aligned)
    add_conventional_kappa_ci(overall, aligned)
    by_source = group_stats(aligned, "review_source", "review_source")
    by_dimension = group_stats(aligned, "dimension", "dimension")
    by_item = group_stats(aligned, "question_id", "question_id")
    by_paper = group_stats(aligned, "paper_code", "paper_code")
    confusion = build_confusion_rows(aligned)
    label_distribution = build_label_distribution_rows(
        aligned,
        annotator_a_metadata["annotator_id"],
        annotator_b_metadata["annotator_id"],
    )
    disagreement_pairs = top_disagreement_pairs(aligned)
    disagreements = [row for row in aligned if not row["exact_agreement"]]

    write_csv(output_dir / "overall_summary.csv", [overall])
    write_csv(output_dir / "agreement_by_review_source.csv", by_source)
    write_csv(output_dir / "agreement_by_dimension.csv", by_dimension)
    write_csv(output_dir / "agreement_by_item.csv", by_item)
    write_csv(output_dir / "agreement_by_paper.csv", by_paper)
    write_csv(output_dir / "confusion_matrix.csv", confusion)
    write_csv(output_dir / "label_distribution.csv", label_distribution)
    write_csv(output_dir / "disagreement_types.csv", disagreement_pairs)
    write_csv(output_dir / "disagreements.csv", disagreements)
    write_csv(output_dir / "aligned_judgements.csv", aligned)
    write_readme(
        output_dir / "README.md",
        overall,
        by_source,
        by_dimension,
        label_distribution,
        annotator_a_metadata,
        annotator_b_metadata,
        disagreement_pairs,
    )

    output_names = [
        "README.md",
        "agreement_by_dimension.csv",
        "agreement_by_item.csv",
        "agreement_by_paper.csv",
        "agreement_by_review_source.csv",
        "aligned_judgements.csv",
        "confusion_matrix.csv",
        "disagreement_types.csv",
        "disagreements.csv",
        "label_distribution.csv",
        "overall_summary.csv",
        "summary.json",
    ]
    summary = {
        "analysis": "Formal human IAA",
        "created_at_utc": datetime.now(timezone.utc).isoformat(),
        "primary_statistic": "unweighted Cohen's kappa over four nominal labels",
        "labels": NOMINAL_LABELS,
        "ordinal_supplement_labels": ORDINAL_LABELS,
        "annotation_source": "P001-P012 paper sheets",
        "analysis_options": {
            "ignore_detailed_judgements": args.ignore_detailed_judgements,
        },
        "overall": overall,
        "annotator_a": annotator_a_metadata,
        "annotator_b": annotator_b_metadata,
        "alignment_checks": {
            "aligned_rows": len(aligned),
            "unique_keys": len(
                {
                    (row["paper_code"], row["review_source"], row["question_id"])
                    for row in aligned
                }
            ),
            "missing_or_invalid_labels": 0,
        },
        "outputs": output_names,
    }
    (output_dir / "summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
